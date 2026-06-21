"""Generate the automotive 'hero' spreadsheet for the MCP server demo.

A deliberately messy regional dealer-group sales/orders workbook, structured the
way real quarterly exports actually look: a merged title, then per-region /
per-month sub-blocks each with their own banner + header, blank separator rows,
sparse cells (VIN only on fulfilled orders, mostly-blank trailing columns), and
inconsistent status casing. Pricing is MSRP-deterministic with a small discount
set, so identical configurations repeat — which is what the inverted-index
encoding collapses, while the blank gaps are what the anchor encoding collapses.
The mess + sparsity + repetition is the point: it's what makes SheetCompressor
show a real token saving while preserving every value (unlike formatAggregation,
which is lossy).

Run:  python examples/build_hero_file.py
Output: examples/northstar-auto-q3-2025.xlsx  (+ prints real token numbers)
"""

import os
import random

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(20260620)  # reproducible — same workbook every run

HEADERS = [
    "Order ID", "Order Date", "Dealership", "Make", "Model", "Trim", "VIN",
    "Qty", "Unit Price", "Discount %", "Total", "Salesperson", "Status",
    # Mostly-blank trailing columns — realistic for a dealer export, and the
    # homogeneous empty regions are exactly what the anchor encoding collapses.
    "Trade-In Value", "Financing", "APR %", "Promo Code", "Notes",
]
NCOLS = len(HEADERS)

# Domestic-leaning lineup (OneMagnify's automotive vertical skews domestic).
# MSRP is fixed per model so identical orders repeat (helps inverted-index).
LINEUP = {
    "Ford":      [("F-150", "Lariat", 58000), ("Explorer", "XLT", 42000),
                  ("Bronco", "Badlands", 49000), ("Escape", "ST-Line", 31000),
                  ("Mustang Mach-E", "Premium", 47000)],
    "Chevrolet": [("Silverado 1500", "LT", 56000), ("Equinox", "RS", 33000),
                  ("Tahoe", "Z71", 64000), ("Malibu", "LS", 27000),
                  ("Blazer EV", "RS", 52000)],
    "Jeep":      [("Grand Cherokee", "Limited", 46000), ("Wrangler", "Rubicon", 44000),
                  ("Compass", "Latitude", 30000)],
    "Ram":       [("1500", "Big Horn", 51000), ("2500", "Laramie", 62000)],
    "GMC":       [("Sierra 1500", "SLT", 57000), ("Acadia", "Denali", 43000)],
    "Cadillac":  [("Lyriq", "Luxury", 61000), ("XT5", "Premium Luxury", 48000)],
}
MAKES = list(LINEUP)

REGIONS = {
    "Midwest": ["Northstar Ford Detroit", "Lakeshore Chevrolet", "Motor City Jeep-Ram", "Heartland GMC"],
    "Southeast": ["Sunbelt Ford Atlanta", "Palmetto Chevrolet", "Gulf Coast Cadillac", "Magnolia Jeep-Ram"],
    "West": ["Pacific Ford", "Sierra Chevrolet", "Mojave GMC", "Cascade Cadillac"],
    "Northeast": ["Liberty Ford Boston", "Empire Chevrolet", "Bay State Jeep-Ram", "Granite Cadillac"],
}
MONTHS = [("2025-07", 7), ("2025-08", 8), ("2025-09", 9)]
SALESPEOPLE = ["A. Okafor", "M. Tran", "J. Delgado", "S. Petrov", "R. Calhoun",
               "L. Nakamura", "B. Whitfield", "C. Okonkwo", "D. Ramirez", "K. Bauer"]
# Inconsistent casing on purpose — repetitive low-cardinality columns are what
# the inverted-index encoding crushes.
STATUSES = ["Delivered", "delivered", "In Transit", "PENDING", "Pending", "Invoiced"]
DISCOUNTS = [0, 0, 0, 5, 10]
FINANCING = ["Lease", "Loan", "Cash"]
PROMOS = ["FALL25", "FLEET10", "LOYALTY", "EVREBATE"]

wb = Workbook()
ws = wb.active
ws.title = "Q3 Sales"

title_font = Font(bold=True, size=14)
region_font = Font(bold=True, size=12, color="FFFFFF")
region_fill = PatternFill("solid", fgColor="1F4E78")
month_font = Font(bold=True, italic=True, color="1F4E78")
header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")

# --- Title block (merged, with a stray subtitle) -----------------------------
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
c = ws.cell(row=1, column=1, value="Northstar Auto Group - Q3 2025 Vehicle Sales & Orders")
c.font = title_font
c.alignment = Alignment(horizontal="left")
ws.cell(row=3, column=2, value="Generated 2025-10-02 | CONFIDENTIAL | figures in USD")

order_seq = 1040
row = 5  # leave blank rows 2 and 4 as mess

for region, dealers in REGIONS.items():
    # Region banner (merged)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    bc = ws.cell(row=row, column=1, value=f"Region: {region}")
    bc.font = region_font
    bc.fill = region_fill
    row += 2  # blank row after region banner

    region_total = 0.0
    for month_label, month_num in MONTHS:
        # Month sub-banner
        ws.cell(row=row, column=1, value=f"{month_label}").font = month_font
        row += 1

        # Header row (repeats per block — repetition the encodings exploit)
        for j, h in enumerate(HEADERS, start=1):
            hc = ws.cell(row=row, column=j, value=h)
            hc.font = header_font
            hc.fill = header_fill
        row += 1

        for _ in range(random.randint(28, 40)):
            make = random.choice(MAKES)
            model, trim, msrp = random.choice(LINEUP[make])
            dealer = random.choice(dealers)
            qty = random.choice([1, 1, 1, 2, 3])  # fleet orders occasionally >1
            unit = msrp                            # MSRP-deterministic → repeats
            status = random.choice(STATUSES)
            discount = random.choice(DISCOUNTS) if random.random() < 0.55 else None
            fulfilled = status.lower() in ("delivered", "invoiced")
            # VIN only on fulfilled orders, and only ~half of those — sparse.
            vin = ("1" + "".join(random.choice("ABCDEFGHJKLMNPRSTUVWXYZ0123456789")
                                 for _ in range(16))) if (fulfilled and random.random() < 0.5) else None
            trim_val = trim if random.random() < 0.85 else None  # occasional blank trim
            eff_disc = discount or 0
            total = round(unit * qty * (1 - eff_disc / 100.0), 2)
            region_total += total

            # Trailing columns: mostly blank on purpose (sparse).
            trade_in = random.choice([6000, 9000, 12000, 18000]) if random.random() < 0.25 else None
            financing = random.choice(FINANCING) if random.random() < 0.40 else None
            apr = 5.9 if (financing == "Loan") else None
            promo = random.choice(PROMOS) if random.random() < 0.15 else None
            note = "fleet account" if (qty >= 2 and random.random() < 0.5) else None

            d = random.randint(1, 28)
            vals = [
                f"NS-{order_seq}", f"{month_label}-{d:02d}", dealer, make, model, trim_val,
                vin, qty, unit, discount, total, random.choice(SALESPEOPLE), status,
                trade_in, financing, apr, promo, note,
            ]
            for j, v in enumerate(vals, start=1):
                if v is not None:
                    ws.cell(row=row, column=j, value=v)
            order_seq += 1
            row += 1

        row += 1  # blank row between months

    # Region subtotal (label under 'Dealership', value under 'Total')
    ws.cell(row=row, column=3, value=f"{region} subtotal").font = Font(bold=True, italic=True)
    ws.cell(row=row, column=11, value=round(region_total, 2)).font = Font(bold=True)
    row += 3  # blank separator before next region

ws.cell(row=row + 1, column=4, value="Prepared by Regional Ops - subject to revision")

for j in range(1, NCOLS + 1):
    ws.column_dimensions[get_column_letter(j)].width = 16

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "northstar-auto-q3-2025.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"rows used: {row}")

# --- Verify it reads + compresses, and print the real demo numbers -----------
from sheet_compressor import compress
from sheet_compressor.adapters.xlsx import read_sheet

grid = read_sheet(out)
result = compress(grid)
raw = result["rawBaseline"]["tokenEstimate"]
print(f"\nraw baseline tokens: {raw:,}")
for enc in ("anchor", "invertedIndex", "formatAggregation"):
    e = result["encodings"][enc]
    t = e["tokenEstimate"]
    keeps = "Silverado" in e["string"]
    print(f"  {enc:18s} {t:>7,} tokens   ({raw / t:>4.1f}x)   values preserved: {keeps}")
