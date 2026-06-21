"""Generate the bundled generic order ledger used by the Resource and the tests.

A simple, predictable order log — fewer rows than the hero file, no merged
title or per-region sub-blocks. Its job is to be a *proven, stable* sample for
the `sheet://examples/{name}` Resource and to give the Seam 1 tests a second
.xlsx (per the issue acceptance criteria) that's easy to verify by hand.

Run:  python examples/build_ledger.py
Output: examples/sample-orders.xlsx
"""

import os
import random

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

random.seed(20260620)

HEADERS = ["Order ID", "Date", "Customer", "SKU", "Qty", "Unit Price", "Total", "Status"]

CUSTOMERS = ["Acme Co", "Globex", "Initech", "Umbrella", "Stark Industries",
             "Wayne Enterprises", "Hooli", "Pied Piper", "Soylent", "Tyrell Corp"]
SKUS = [
    ("SKU-001", "Widget A", 19.99),
    ("SKU-002", "Widget B", 29.99),
    ("SKU-003", "Gizmo X", 49.50),
    ("SKU-004", "Gizmo Y", 79.00),
    ("SKU-005", "Sprocket M", 12.25),
]
STATUSES = ["Shipped", "shipped", "Pending", "PENDING", "Refunded"]

wb = Workbook()
ws = wb.active
ws.title = "Orders"

# Header row, lightly styled.
for col, value in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=value)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

# 120 rows is plenty to show anchor's structural skeleton without bloating the file.
for i in range(120):
    order_id = f"O-{1000 + i:05d}"
    customer = random.choice(CUSTOMERS)
    sku_id, _, unit_price = random.choice(SKUS)
    qty = random.randint(1, 8)
    total = round(qty * unit_price, 2)
    status = random.choice(STATUSES)
    date = f"2025-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}"
    ws.append([order_id, date, customer, sku_id, qty, unit_price, total, status])

OUT = os.path.join(os.path.dirname(__file__), "sample-orders.xlsx")
wb.save(OUT)
print(f"Wrote {OUT}")
